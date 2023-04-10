# 这个文件记录了爬虫所用到的一些必要函数

import requests
import re
from bs4 import BeautifulSoup
import openpyxl as op



# 初始化爬虫，运用beautifulsoup库返回一个soup
def first(header,urls,timeouts):
    res = requests.get(urls, headers=header,timeout=timeouts)
    res.encoding = 'gbk'
    soup = BeautifulSoup(res.text,'html.parser')
    return soup


# 返回所有要爬取的连接与月份名字
def link_obtain(soup):
    #filename = (soup.title.string.replace('\r','').replace('\n','').replace('\t','')).split('_')[2]
    hreflist = soup.find_all('div',class_='months')
    hreflist = str(hreflist).split('</a>')
    link = []
    other_name = []
    for i in range(len(hreflist)):
        bb = hreflist[i]
        link.append(bb[bb.find('href=') + 6:bb.find('title') -2])
        other_name.append(bb[bb.find('title') + 7:bb.find('天气') + 2])
    return link,other_name





# 关键点的提取
def key_obtain(soup):
    tem_list = soup.find_all('tr')  
    return tem_list


# 每个城市数据文件的表头写入
def table_tag_write(tem_list,filename):
    wb = op.Workbook() # 创建工作薄对象
    ws = wb['Sheet'] # 创建子表 

    # 表头提取出来了
    table1 = re.findall(r'<td>\n[\s]*<b>(.*?)</b></td>',str(tem_list[0]))
    for oo in range(len(table1)):
        ws.cell(row=1,column = oo + 1).value = table1[oo] 
    wb.save(filename + '.xlsx') 
    return wb,ws

# 数据文件中间部分的写入
def table_body_write(tem_list,local,filename):
    wb = op.load_workbook(filename + '.xlsx') # 打开MY_EXCEL.xlsx文件
    ws = wb.active # 激活工作区
    for i in range(1,len(tem_list)):
        time = str(tem_list[i].a.string).replace(' ','').replace('\n','')
        aa = (str(tem_list[i]).replace(' ','')).split("<td>")
        for j in range(len(aa)):
            aa[j] = aa[j].replace('\r','')
            aa[j] = aa[j].replace('\n','')
            aa[j] = aa[j].replace('</td>','')
            aa[j] = aa[j].replace('</tr>','')
        aa[1] = time
        aa = aa[1:]
        for k in range(len(aa)):
            ws.cell(row=local + i + 1,column=k+1).value = aa[k] # 将数据data写入excel中的第i行第j列
    local = local + len(tem_list)-1
    wb.save(filename + '.xlsx') # 保存excel表
    return local











